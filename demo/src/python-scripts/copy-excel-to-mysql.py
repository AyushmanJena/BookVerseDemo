import openpyxl
import pymysql

# Database connection parameters
db_config = {
    'host': 'localhost',          # replace with your MySQL host if different
    'user': 'springstudent',      # replace with your MySQL username
    'password': 'springstudent',  # replace with your MySQL password
    'database': 'book_reviews',   # replace with your database name
}

# Load the Excel file
excel_file_path = "D:/Book Demo/demo/src/python-scripts/bookverse.xlsx"  # replace with your Excel file path
workbook = openpyxl.load_workbook(excel_file_path) 
sheet = workbook.active

# Establish a connection to the MySQL database
connection = pymysql.connect(**db_config)
cursor = connection.cursor()

# Generate recommendations and insert into book table
recommendations_data = []

for row in sheet.iter_rows(min_row=2, values_only=True):
    book_id = row[0]  # Manually set book_id from the first column
    if book_id is not None and book_id != '':
        recommendations_data.append((
            book_id,    # Valid book_id
            4,          # Set book_rating to 4 for all books
            row[1],     # NAME OF THE AUTHOR
            row[2],     # NAME OF THE BOOK
            row[3],     # IMAGE URL
            None, None, None, None, None  # Leave similar book columns empty (NULL)
        ))
    else:
        print(f"Skipping row {row}: book_id is empty or None.")

# Insert into the book table only if recommendations_data is not empty
if recommendations_data:
    book_insert_query = """
    INSERT INTO book (book_id, book_rating, book_author, book_title, book_cover_link, 
                      similar_1, similar_2, similar_3, similar_4, similar_5)
    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """
    
    # Execute batch insert into book table
    cursor.executemany(book_insert_query, recommendations_data)
    connection.commit()
else:
    print("No valid book data to insert into the book table.")

# Insert data into the book_reviews.model table if recommendations_data is not empty
model_data = []

for row in sheet.iter_rows(min_row=2, values_only=True):
    book_id = row[0]  # Get book_id from the first column
    if book_id is not None and book_id != '':
        model_data.append((
            book_id,  # BOOK ID
            row[4],   # HUMOUR
            row[5],   # COMEDY
            row[6],   # EDUCATIONAL
            row[7],   # THRILLER
            row[8],   # ROMANCE
            row[9],   # PHILOSOPHICAL
            row[10],  # MOTIVATIONAL
            row[11],  # FINANCE
            row[12],  # DRAMA
            row[13],  # HORROR
            row[14],  # BIOPIC
            row[15],  # FICTION
            row[16],  # MYTHOLOGY
            row[17],  # SCI-FI
            row[18],  # ADVENTURE
            row[19],  # COMICS
            row[20]   # RELIGIOUS TEXT
        ))

# Execute batch insert into model table only if model_data is not empty
if model_data:
    model_insert_query = """
    INSERT INTO model (book_id, humour, comedy, education, thriller, romance, 
                       philosophical, motivational, finance, drama, horror, 
                       biopic, fiction, mythology, sci_fi, adventure, comics, religious_text)
    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """
    
    cursor.executemany(model_insert_query, model_data)
    connection.commit()
else:
    print("No valid model data to insert into the model table.")

# Close cursor and connection
cursor.close()
connection.close()

print("Data successfully inserted into both tables.")
